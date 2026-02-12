"""Parser for e.on XLSX reports exported from eBOK."""

import openpyxl
from typing import List, Dict, Any, Optional


def parse_eon_xlsx(filepath: str) -> List[Dict[str, Any]]:
    """Parse e.on XLSX report and extract document records.
    
    Returns list of dicts with keys:
        doc_type, doc_number, issue_date, due_date, amount_pln,
        payment_status, location, account_number
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        vals = [cell.value for cell in row]
        if vals and vals[0] == "No.":
            header_row = row_idx
            break

    if header_row is None:
        raise ValueError("Could not find header row in e.on XLSX")

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        no, account, doc_type, doc_number, issue_date, due_date, amount, status, *rest = row
        if no is None or doc_type is None:
            continue

        # Extract location from account string like "80000080441 Płatnicza 65"
        location = ""
        if account and isinstance(account, str):
            parts = account.split(" ", 1)
            if len(parts) > 1:
                location = parts[1].strip()

        record = {
            "provider": "eon",
            "doc_type": _normalize_doc_type(doc_type),
            "doc_type_original": doc_type,
            "doc_number": str(doc_number) if doc_number else None,
            "issue_date": _normalize_date(issue_date),
            "due_date": _normalize_date(due_date),
            "amount_pln": float(amount) if amount and amount != "-" else None,
            "payment_status": status if status and status != "-" else None,
            "location": location,
            "account_number": account.split(" ", 1)[0] if account and isinstance(account, str) else None,
            "utility_type": "electricity",
        }
        records.append(record)

    return records


def extract_eon_consumption(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Extract consumption records from e.on data.
    
    Note: e.on XLSX report only has amounts, not detailed consumption.
    For detailed kWh data we'd need the actual invoice PDFs.
    We create consumption records from 'faktura_rozliczeniowa' entries
    as they represent actual metered usage periods.
    """
    consumption = []
    for rec in records:
        if rec["doc_type"] not in ("faktura_rozliczeniowa", "prognoza"):
            continue
        if rec["amount_pln"] is None:
            continue

        consumption.append({
            "provider": "eon",
            "utility_type": "electricity",
            "location": rec["location"],
            "period_start": rec.get("due_date", rec["issue_date"]),
            "period_end": rec["issue_date"],
            "cost_gross": rec["amount_pln"],
            "is_estimate": 1 if rec["doc_type"] == "prognoza" else 0,
            "doc_number": rec["doc_number"],
            "doc_type": rec["doc_type"],
        })

    return consumption


def _normalize_doc_type(doc_type: str) -> str:
    mapping = {
        "Faktura rozliczeniowa": "faktura_rozliczeniowa",
        "Prognoza zużycia": "prognoza",
        "Nota odsetkowa": "nota_odsetkowa",
        "Wpłata bankowa": "wplata",
    }
    return mapping.get(doc_type, doc_type.lower().replace(" ", "_"))


def _normalize_date(date_val) -> Optional[str]:
    if date_val is None or date_val == "-":
        return None
    if isinstance(date_val, str):
        # Convert DD-MM-YYYY to YYYY-MM-DD
        parts = date_val.split("-")
        if len(parts) == 3 and len(parts[0]) == 2:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
        return date_val
    return str(date_val)
